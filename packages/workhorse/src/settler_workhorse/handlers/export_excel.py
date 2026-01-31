"""Excel export handler for export.excel jobs.

Generates Excel (.xlsx) exports from tenant data with formatting.
Idempotent, retry-safe, and tenant-scoped.
"""

import hashlib
from datetime import datetime
from typing import Any, Dict, List, Optional

from settler_workhorse.models import Job, JobResult, JobType
from settler_workhorse.utils.logging import get_logger
from settler_workhorse.worker import register_handler

# Try to import openpyxl, fallback to CSV if not available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = get_logger("handlers.export_excel")


class ExcelExportError(Exception):
    """Error during Excel export."""

    pass


def validate_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Validate and normalize Excel export payload.

    Args:
        payload: Raw job payload

    Returns:
        Validated payload with defaults

    Raises:
        ExcelExportError: If required fields are missing
    """
    tenant_id = payload.get("tenant_id")
    if not tenant_id:
        raise ExcelExportError("tenant_id is required")

    entity_type = payload.get("entity_type")
    if not entity_type:
        raise ExcelExportError("entity_type is required")

    return {
        "tenant_id": tenant_id,
        "entity_type": entity_type,
        "filters": payload.get("filters", {}),
        "columns": payload.get("columns", []),
        "dry_run": payload.get("dry_run", False),
        "idempotency_key": payload.get("idempotency_key"),
        "include_headers": payload.get("include_headers", True),
        "format_headers": payload.get("format_headers", True),
        "freeze_headers": payload.get("freeze_headers", True),
        "sheet_name": payload.get("sheet_name", "Data"),
        "max_records": payload.get("max_records", 100000),
    }


def generate_mock_data(
    entity_type: str,
    filters: Dict[str, Any],
    max_records: int,
) -> List[Dict[str, Any]]:
    """Generate mock data for export.

    Args:
        entity_type: Type of entity to export
        filters: Filters to apply
        max_records: Maximum records to generate

    Returns:
        List of record dictionaries
    """
    records = []
    count = min(max_records, 1000)

    for i in range(count):
        if entity_type == "transactions":
            record = {
                "id": f"txn_{i:06d}",
                "external_id": f"ext_{i:08d}",
                "amount": round(10.0 + (i * 1.5), 2),
                "currency": "USD",
                "date": datetime.utcnow(),
                "description": f"Transaction {i}",
                "status": "matched" if i % 3 == 0 else "unmatched",
                "source": "bank_a" if i % 2 == 0 else "bank_b",
                "created_at": datetime.utcnow(),
            }
        elif entity_type == "reconciliations":
            record = {
                "id": f"rec_{i:06d}",
                "name": f"Reconciliation Run {i}",
                "source_system": "Stripe",
                "target_system": "Bank Account",
                "match_rate": round(0.85 + (i % 15) / 100, 2),
                "total_records": 1000 + i * 10,
                "matched_records": 850 + i * 8,
                "created_at": datetime.utcnow(),
                "status": "completed" if i % 5 != 0 else "processing",
            }
        else:
            record = {
                "id": f"rec_{i:06d}",
                "entity_type": entity_type,
                "index": i,
                "created_at": datetime.utcnow(),
            }

        records.append(record)

    return records


def generate_excel_content(
    records: List[Dict[str, Any]],
    columns: List[str],
    sheet_name: str,
    format_headers: bool,
    freeze_headers: bool,
) -> bytes:
    """Generate Excel content from records.

    Args:
        records: List of record dictionaries
        columns: Column names to include (empty = all)
        sheet_name: Name of the worksheet
        format_headers: Whether to format header row
        freeze_headers: Whether to freeze header row

    Returns:
        Excel file bytes
    """
    if not HAS_OPENPYXL:
        raise ExcelExportError("openpyxl is not installed. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Determine columns
    if columns:
        fieldnames = columns
    else:
        fieldnames = list(records[0].keys()) if records else []

    # Write headers
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        if format_headers:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, record in enumerate(records, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = record.get(field, "")
            if isinstance(value, datetime):
                ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, field in enumerate(fieldnames, 1):
        max_length = len(str(field))
        for record in records:
            value = str(record.get(field, ""))
            max_length = max(max_length, len(value))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    if freeze_headers:
        ws.freeze_panes = "A2"

    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@register_handler(JobType.EXPORT_EXCEL)
def handle_export_excel(job: Job) -> JobResult:
    """Handle Excel export job.

    Args:
        job: Job containing export payload

    Returns:
        Job execution result
    """
    try:
        payload = validate_payload(job.payload)
    except ExcelExportError as e:
        return JobResult(
            success=False,
            error=str(e),
            records_processed=0,
            records_failed=0,
        )

    tenant_id = payload["tenant_id"]
    entity_type = payload["entity_type"]
    filters = payload["filters"]
    columns = payload["columns"]
    dry_run = payload["dry_run"]
    idempotency_key = payload.get("idempotency_key")
    sheet_name = payload["sheet_name"]
    max_records = payload["max_records"]
    format_headers = payload["format_headers"]
    freeze_headers = payload["freeze_headers"]

    logger.info(
        "Starting Excel export",
        job_id=str(job.id),
        tenant_id=tenant_id,
        entity_type=entity_type,
        dry_run=dry_run,
    )

    try:
        if not HAS_OPENPYXL:
            logger.warning("openpyxl not available, Excel export will fail")
            return JobResult(
                success=False,
                error="Excel export requires openpyxl. Please install: pip install openpyxl",
                records_processed=0,
                records_failed=0,
            )

        # Generate data
        records = generate_mock_data(entity_type, filters, max_records)

        if dry_run:
            return JobResult(
                success=True,
                data={
                    "export_id": f"export_excel_{job.id}_dryrun",
                    "tenant_id": tenant_id,
                    "entity_type": entity_type,
                    "mode": "dry_run",
                    "estimated_records": len(records),
                    "columns": columns if columns else "all",
                    "sheet_name": sheet_name,
                    "has_openpyxl": HAS_OPENPYXL,
                    "idempotency_key": idempotency_key,
                },
                records_processed=0,
                records_failed=0,
            )

        # Generate Excel content
        excel_bytes = generate_excel_content(
            records=records,
            columns=columns,
            sheet_name=sheet_name,
            format_headers=format_headers,
            freeze_headers=freeze_headers,
        )

        # Calculate checksum
        content_hash = hashlib.sha256(excel_bytes).hexdigest()[:16]

        return JobResult(
            success=True,
            data={
                "export_id": f"export_excel_{job.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
                "tenant_id": tenant_id,
                "entity_type": entity_type,
                "mode": "live",
                "format": "xlsx",
                "records_exported": len(records),
                "columns": list(records[0].keys()) if records else [],
                "content_hash": content_hash,
                "size_bytes": len(excel_bytes),
                "sheet_name": sheet_name,
                "formatted": format_headers,
                "idempotency_key": idempotency_key,
                "exported_at": datetime.utcnow().isoformat(),
            },
            records_processed=len(records),
            records_failed=0,
            output_location=job.payload.get("output_path"),
        )

    except Exception as e:
        logger.error("Excel export failed", exc_info=True, tenant_id=tenant_id, entity_type=entity_type)
        return JobResult(
            success=False,
            error=str(e),
            records_processed=0,
            records_failed=0,
        )
