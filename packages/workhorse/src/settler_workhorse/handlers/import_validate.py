"""Import validation handler for import.validate jobs.

Validates CSV/Excel files before actual import.
Idempotent, retry-safe, and tenant-scoped.
"""

import csv
import hashlib
import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from settler_workhorse.models import Job, JobResult, JobType
from settler_workhorse.utils.logging import get_logger
from settler_workhorse.worker import register_handler

# Try to import optional dependencies
try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = get_logger("handlers.import_validate")


class ImportValidationError(Exception):
    """Error during import validation."""

    pass


def validate_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Validate and normalize import validation payload.

    Args:
        payload: Raw job payload

    Returns:
        Validated payload with defaults

    Raises:
        ImportValidationError: If required fields are missing
    """
    tenant_id = payload.get("tenant_id")
    if not tenant_id:
        raise ImportValidationError("tenant_id is required")

    file_content_base64 = payload.get("file_content_base64")
    if not file_content_base64:
        raise ImportValidationError("file_content_base64 is required")

    return {
        "tenant_id": tenant_id,
        "file_content_base64": file_content_base64,
        "import_type": payload.get("import_type", "csv"),  # csv, xlsx
        "expected_columns": payload.get("expected_columns", []),
        "required_columns": payload.get("required_columns", []),
        "dry_run": payload.get("dry_run", False),
        "idempotency_key": payload.get("idempotency_key"),
        "max_preview_rows": payload.get("max_preview_rows", 10),
        "column_mapping": payload.get("column_mapping", {}),
        "validation_rules": payload.get("validation_rules", {}),
    }


def detect_encoding(content_bytes: bytes) -> str:
    """Detect file encoding.

    Args:
        content_bytes: Raw file content

    Returns:
        Detected encoding
    """
    if HAS_CHARDET:
        result = chardet.detect(content_bytes)
        return result.get("encoding", "utf-8") or "utf-8"
    return "utf-8"


def validate_csv_content(
    content_bytes: bytes,
    expected_columns: List[str],
    required_columns: List[str],
    max_preview_rows: int,
) -> Tuple[bool, List[Dict[str, Any]], Dict[str, Any]]:
    """Validate CSV content.

    Args:
        content_bytes: Raw CSV bytes
        expected_columns: Expected column names
        required_columns: Required column names
        max_preview_rows: Maximum rows for preview

    Returns:
        Tuple of (is_valid, preview_rows, statistics)
    """
    errors = []
    preview_rows = []
    stats = {
        "total_rows": 0,
        "valid_rows": 0,
        "error_rows": 0,
        "missing_required": [],
    }

    try:
        encoding = detect_encoding(content_bytes)
        content_str = content_bytes.decode(encoding)

        reader = csv.DictReader(io.StringIO(content_str))
        actual_columns = reader.fieldnames or []

        # Check for expected columns
        if expected_columns:
            missing = [col for col in expected_columns if col not in actual_columns]
            if missing:
                errors.append(f"Missing expected columns: {', '.join(missing)}")

        # Check for required columns
        missing_required = [col for col in required_columns if col not in actual_columns]
        if missing_required:
            errors.append(f"Missing required columns: {', '.join(missing_required)}")
            stats["missing_required"] = missing_required

        # Validate rows
        for idx, row in enumerate(reader, start=2):  # Start at 2 (header is 1)
            stats["total_rows"] += 1
            row_errors = []

            # Check required fields
            for col in required_columns:
                if col in row and not row[col]:
                    row_errors.append(f"Empty value in required column '{col}'")

            if row_errors:
                stats["error_rows"] += 1
            else:
                stats["valid_rows"] += 1

            # Add to preview
            if len(preview_rows) < max_preview_rows:
                preview_rows.append({
                    "row_number": idx,
                    "data": dict(row),
                    "errors": row_errors,
                })

        is_valid = len(errors) == 0 and stats["error_rows"] == 0

        return is_valid, preview_rows, stats

    except Exception as e:
        return False, [], {"error": str(e), "total_rows": 0, "valid_rows": 0, "error_rows": 0}


def validate_excel_content(
    content_bytes: bytes,
    expected_columns: List[str],
    required_columns: List[str],
    max_preview_rows: int,
) -> Tuple[bool, List[Dict[str, Any]], Dict[str, Any]]:
    """Validate Excel content.

    Args:
        content_bytes: Raw Excel bytes
        expected_columns: Expected column names
        required_columns: Required column names
        max_preview_rows: Maximum rows for preview

    Returns:
        Tuple of (is_valid, preview_rows, statistics)
    """
    if not HAS_OPENPYXL:
        return False, [], {"error": "openpyxl not installed", "total_rows": 0, "valid_rows": 0, "error_rows": 0}

    errors = []
    preview_rows = []
    stats = {
        "total_rows": 0,
        "valid_rows": 0,
        "error_rows": 0,
        "missing_required": [],
    }

    try:
        wb = load_workbook(io.BytesIO(content_bytes))
        ws = wb.active

        # Get headers from first row
        headers = [str(cell.value) if cell.value else f"Column_{i}" for i, cell in enumerate(ws[1], 1)]

        # Check for expected columns
        if expected_columns:
            missing = [col for col in expected_columns if col not in headers]
            if missing:
                errors.append(f"Missing expected columns: {', '.join(missing)}")

        # Check for required columns
        missing_required = [col for col in required_columns if col not in headers]
        if missing_required:
            errors.append(f"Missing required columns: {', '.join(missing_required)}")
            stats["missing_required"] = missing_required

        # Validate rows
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            stats["total_rows"] += 1
            row_data = dict(zip(headers, row))
            row_errors = []

            # Check required fields
            for col in required_columns:
                if col in row_data and not row_data[col]:
                    row_errors.append(f"Empty value in required column '{col}'")

            if row_errors:
                stats["error_rows"] += 1
            else:
                stats["valid_rows"] += 1

            # Add to preview
            if len(preview_rows) < max_preview_rows:
                preview_rows.append({
                    "row_number": idx,
                    "data": row_data,
                    "errors": row_errors,
                })

        is_valid = len(errors) == 0 and stats["error_rows"] == 0

        return is_valid, preview_rows, stats

    except Exception as e:
        return False, [], {"error": str(e), "total_rows": 0, "valid_rows": 0, "error_rows": 0}


@register_handler(JobType.IMPORT_VALIDATE)
def handle_import_validate(job: Job) -> JobResult:
    """Handle import validation job.

    Args:
        job: Job containing validation payload

    Returns:
        Job execution result
    """
    try:
        payload = validate_payload(job.payload)
    except ImportValidationError as e:
        return JobResult(
            success=False,
            error=str(e),
            records_processed=0,
            records_failed=0,
        )

    tenant_id = payload["tenant_id"]
    file_content_base64 = payload["file_content_base64"]
    import_type = payload["import_type"]
    expected_columns = payload["expected_columns"]
    required_columns = payload["required_columns"]
    dry_run = payload["dry_run"]
    idempotency_key = payload.get("idempotency_key")
    max_preview_rows = payload["max_preview_rows"]
    column_mapping = payload["column_mapping"]

    logger.info(
        "Starting import validation",
        job_id=str(job.id),
        tenant_id=tenant_id,
        import_type=import_type,
        dry_run=dry_run,
    )

    try:
        import base64

        # Decode base64 content
        try:
            content_bytes = base64.b64decode(file_content_base64)
        except Exception:
            return JobResult(
                success=False,
                error="Invalid base64 content",
                records_processed=0,
                records_failed=0,
            )

        # Calculate checksum
        content_hash = hashlib.sha256(content_bytes).hexdigest()[:16]

        if dry_run:
            return JobResult(
                success=True,
                data={
                    "validation_id": f"import_val_{job.id}_dryrun",
                    "tenant_id": tenant_id,
                    "import_type": import_type,
                    "mode": "dry_run",
                    "size_bytes": len(content_bytes),
                    "content_hash": content_hash,
                    "expected_columns": expected_columns,
                    "required_columns": required_columns,
                    "idempotency_key": idempotency_key,
                },
                records_processed=0,
                records_failed=0,
            )

        # Validate based on import type
        if import_type == "csv":
            is_valid, preview_rows, stats = validate_csv_content(
                content_bytes, expected_columns, required_columns, max_preview_rows
            )
        elif import_type == "xlsx":
            is_valid, preview_rows, stats = validate_excel_content(
                content_bytes, expected_columns, required_columns, max_preview_rows
            )
        else:
            return JobResult(
                success=False,
                error=f"Unsupported import type: {import_type}",
                records_processed=0,
                records_failed=0,
            )

        return JobResult(
            success=True,
            data={
                "validation_id": f"import_val_{job.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
                "tenant_id": tenant_id,
                "import_type": import_type,
                "mode": "live",
                "is_valid": is_valid,
                "content_hash": content_hash,
                "size_bytes": len(content_bytes),
                "statistics": stats,
                "preview_rows": preview_rows,
                "column_mapping": column_mapping,
                "idempotency_key": idempotency_key,
                "validated_at": datetime.utcnow().isoformat(),
            },
            records_processed=stats.get("total_rows", 0),
            records_failed=stats.get("error_rows", 0),
            output_location=None,
        )

    except Exception as e:
        logger.error("Import validation failed", exc_info=True, tenant_id=tenant_id, import_type=import_type)
        return JobResult(
            success=False,
            error=str(e),
            records_processed=0,
            records_failed=0,
        )
